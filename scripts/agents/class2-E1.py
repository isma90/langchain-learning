import json
import logging
from datetime import datetime

from rich import print as rprint
from rich.pretty import Pretty
from rich.table import Table

from langchain.agents import create_agent
from langchain.chat_models import init_chat_model

from langgraph.checkpoint.memory import InMemorySaver

from pydantic import BaseModel, Field
import pandas as pd
from dotenv import load_dotenv

from scripts.agents.tools.tools import buscar_informacion_nutricional, extraer_datos_nutricionales, \
    ToolMonitoringMiddleware

load_dotenv()

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# ============================================================================
# 1. MODELOS PYDANTIC PARA ESTRUCTURA DE SALIDA
# ============================================================================

class IngredienteNutricional(BaseModel):
    """Representa un ingrediente con su información nutricional"""
    ingrediente: str = Field(..., description="Nombre del ingrediente")
    cantidad: str = Field(..., description="Cantidad del ingrediente (ej: 100g, 1 taza)")
    calorias: float = Field(..., description="Calorías por porción")
    grasas_gramos: float = Field(..., description="Grasas en gramos por porción")


class TotalesNutricionales(BaseModel):
    """Resumen totales de información nutricional"""
    total_calorias: float = Field(..., description="Suma total de calorías")
    total_grasas: float = Field(..., description="Suma total de grasas en gramos")
    cantidad_ingredientes: int = Field(..., description="Cantidad de ingredientes procesados")


class AnálisisReceta(BaseModel):
    """Estructura completa del análisis nutricional de una receta"""
    nombre_receta: str = Field(..., description="Nombre de la receta")
    ingredientes: list[IngredienteNutricional] = Field(
        ..., description="Lista de ingredientes con información nutricional"
    )
    totales: TotalesNutricionales = Field(..., description="Totales nutricionales")
    timestamp: str = Field(
        default_factory=lambda: datetime.now().isoformat(),
        description="Fecha y hora del análisis"
    )

# ============================================================================
# 4. CONFIGURACIÓN DEL AGENTE
# ============================================================================

model = init_chat_model("openai:gpt-4o-mini", temperature=0)

agent = create_agent(
    model=model,
    tools=[buscar_informacion_nutricional, extraer_datos_nutricionales],
    system_prompt=(
        "Eres un asistente especializado en análisis nutricional de recetas. "
        "Tu tarea es:\n"
        "1. Recibir una receta con ingredientes y cantidades\n"
        "2. Para CADA ingrediente, usa buscar_informacion_nutricional para obtener datos\n"
        "3. Analiza los resultados de búsqueda y EXTRAE manualmente (con tu comprensión de lenguaje):\n"
        "   - Calorías por 100g\n"
        "   - Gramos de grasas por 100g\n"
        "   IMPORTANTE: No intentes usar extraer_datos_nutricionales si puedes leer los valores directamente\n"
        "4. Calcula los valores para la cantidad específica de la receta\n"
        "5. Retorna SOLAMENTE un JSON válido con esta estructura exacta:\n"
        "{\n"
        '  "nombre_receta": "nombre de la receta",\n'
        '  "ingredientes": [\n'
        '    {"ingrediente": "nombre", "cantidad": "cantidad", "calorias": número, "grasas_gramos": número},\n'
        "    ...\n"
        "  ],\n"
        '  "totales": {\n'
        '    "total_calorias": suma,\n'
        '    "total_grasas": suma,\n'
        '    "cantidad_ingredientes": número\n'
        "  }\n"
        "}\n"
        "\nNOTA: Usa tu capacidad de lectura y comprensión para extraer los datos. "
        "Si la herramienta no encuenta resultados, usa valores estándar para ese ingrediente."
    ),
    middleware=[ToolMonitoringMiddleware],
    checkpointer=InMemorySaver(),
)


# ============================================================================
# 5. FUNCIÓN PARA GENERAR EXCEL
# ============================================================================

def generar_excel_desde_analisis(analisis: AnálisisReceta, nombre_archivo: str = None) -> str:
    """
    Genera un archivo Excel con los datos nutricionales de la receta.

    Estructura:
    - Hoja "Análisis": tabla con ingredientes y datos nutricionales
    - Fila de totales al final
    - Información del timestamp
    """
    if nombre_archivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"analisis_receta_{timestamp}.xlsx"

    # Preparar datos para DataFrame
    datos_ingredientes = []
    for ing in analisis.ingredientes:
        datos_ingredientes.append({
            "Ingrediente": ing.ingrediente,
            "Cantidad": ing.cantidad,
            "Calorías": ing.calorias,
            "Grasas (g)": ing.grasas_gramos,
        })

    # Agregar fila de totales
    datos_ingredientes.append({
        "Ingrediente": "TOTAL",
        "Cantidad": "",
        "Calorías": analisis.totales.total_calorias,
        "Grasas (g)": analisis.totales.total_grasas,
    })

    df = pd.DataFrame(datos_ingredientes)

    # Crear archivo Excel con formato
    with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
        # Escribir hoja principal
        df.to_excel(writer, sheet_name="Análisis", index=False)

        # Aplicar formato a la hoja
        worksheet = writer.sheets["Análisis"]

        # Ancho de columnas
        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 15
        worksheet.column_dimensions["C"].width = 12
        worksheet.column_dimensions["D"].width = 12

        # Estilos para encabezados (opcional pero mejora presentación)
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Formato para fila de totales (última fila)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True)

        last_row = len(datos_ingredientes) + 1
        for cell in worksheet[last_row]:
            cell.fill = total_fill
            cell.font = total_font

        # Centrar valores numéricos
        for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")

    logger.info(f"✅ Archivo Excel generado: {nombre_archivo}")
    return nombre_archivo


# ============================================================================
# 6. FUNCIÓN PRINCIPAL PARA PROCESAR UNA RECETA
# ============================================================================

def procesar_receta(receta_texto: str, guardar_excel: bool = True) -> AnálisisReceta:
    """
    Procesa una receta usando el agente LangChain y retorna un análisis estructurado.

    Args:
        receta_texto: Texto descriptivo de la receta con ingredientes
        guardar_excel: Si es True, genera un archivo Excel con los resultados

    Returns:
        AnálisisReceta: Objeto Pydantic con los datos nutricionales
    """
    rprint("\n[bold blue]🍳 INICIANDO ANÁLISIS NUTRICIONAL DE RECETA[/bold blue]\n")
    rprint(f"[italic]Receta:[/italic] {receta_texto[:100]}...\n")

    # Extraer ingredientes de la receta para asegurar que procese todos
    lineas = receta_texto.strip().split('\n')
    nombre_receta_raw = lineas[0].split(':')[0].strip() if ':' in lineas[0] else "Receta"

    ingredientes_raw = []
    for linea in lineas[1:]:
        linea = linea.strip()
        if linea and linea.startswith('-'):
            ingredientes_raw.append(linea[1:].strip())

    # Crear lista numerada para que el agente sea más explícito
    lista_ingredientes = "\n".join([f"{i+1}. {ing}" for i, ing in enumerate(ingredientes_raw)])

    # Ejecutar agente
    lista_nombres = ", ".join([f'"{ing}"' for ing in ingredientes_raw])
    mensaje_usuario = f"""TAREA CRÍTICA: Analiza esta receta y busca información nutricional para CADA UNO DE LOS INGREDIENTES.

NOMBRE DE RECETA: {nombre_receta_raw}

INGREDIENTES A PROCESAR - LISTA COMPLETA ({len(ingredientes_raw)} total):
{lista_ingredientes}

INSTRUCCIONES ESTRICTAS:
1. Busca información nutricional para CADA UNO de estos {len(ingredientes_raw)} ingredientes
2. Para cada ingrediente:
   - Usa la herramienta buscar_informacion_nutricional con el nombre exacto
   - Lee los resultados de búsqueda
   - Extrae: calorías por 100g y grasas en gramos por 100g
   - Calcula para la cantidad exacta en la receta
3. Crea entrada en array para CADA ingrediente con su NOMBRE REAL (no "ingrediente1")
4. Suma TODOS para los totales

FORMATO DE RESPUESTA - RETORNA SOLAMENTE ESTE JSON (con TODOS los {len(ingredientes_raw)} ingredientes):
{{
  "nombre_receta": "{nombre_receta_raw}",
  "ingredientes": [
    {{"ingrediente": "nombre_real_1", "cantidad": "cantidad_1", "calorias": numero, "grasas_gramos": numero}},
    {{"ingrediente": "nombre_real_2", "cantidad": "cantidad_2", "calorias": numero, "grasas_gramos": numero}},
    {{"ingrediente": "nombre_real_3", "cantidad": "cantidad_3", "calorias": numero, "grasas_gramos": numero}},
    ... CONTINÚA CON TODOS LOS {len(ingredientes_raw)} INGREDIENTES ...
    {{"ingrediente": "nombre_real_{len(ingredientes_raw)}", "cantidad": "cantidad_{len(ingredientes_raw)}", "calorias": numero, "grasas_gramos": numero}}
  ],
  "totales": {{
    "total_calorias": SUMA_TODAS_CALORIAS,
    "total_grasas": SUMA_TODAS_GRASAS,
    "cantidad_ingredientes": {len(ingredientes_raw)}
  }}
}}

⚠️ REQUISITOS CRÍTICOS:
✓ TODOS los {len(ingredientes_raw)} ingredientes DEBEN estar en la lista con sus NOMBRES REALES
✓ cantidad_ingredientes DEBE ser exactamente {len(ingredientes_raw)}
✓ Retorna SOLO el JSON válido (sin explicaciones antes o después)
✓ Todos los valores numéricos DEBEN ser números, no strings
✓ Los totales deben ser la suma exacta
✓ Nombres de ingredientes a usar: {lista_nombres}
"""

    thread_id = f"receta-{datetime.now().timestamp()}"

    resultado = agent.invoke(
        {"messages": [{"role": "user", "content": mensaje_usuario}]},
        config={"configurable": {"thread_id": thread_id}},
    )

    # Extraer respuesta del agente - buscar JSON en el contenido
    import re
    respuesta_final = None

    for msg in resultado["messages"]:
        if hasattr(msg, "content") and msg.content:
            contenido = msg.content

            # Intentar parsear directamente como JSON completo
            try:
                json_match = json.loads(contenido)
                if isinstance(json_match, dict) and "nombre_receta" in json_match:
                    respuesta_final = json_match
                    logger.info("✅ JSON extraído directamente del contenido")
                    break
            except:
                pass

            # Si no funciona, buscar JSON dentro del contenido (entre { })
            if not respuesta_final:
                try:
                    # Buscar el primer { y el último }
                    inicio = contenido.find('{')
                    fin = contenido.rfind('}')
                    if inicio != -1 and fin != -1 and fin > inicio:
                        json_str = contenido[inicio:fin+1]
                        json_match = json.loads(json_str)
                        if isinstance(json_match, dict) and "nombre_receta" in json_match:
                            respuesta_final = json_match
                            logger.info("✅ JSON extraído del contenido con delimitadores")
                            break
                except Exception as e:
                    logger.debug(f"No se encontró JSON en el mensaje: {e}")
                    pass

    if not respuesta_final:
        # Fallback: construir análisis minimalista
        logger.warning("⚠️ No se pudo extraer JSON estructurado del agente")
        respuesta_final = {
            "nombre_receta": "Receta",
            "ingredientes": [],
            "totales": {"total_calorias": 0, "total_grasas": 0, "cantidad_ingredientes": 0}
        }
        logger.info(f"Mensajes recibidos del agente: {len(resultado.get('messages', []))} mensajes")
        for i, msg in enumerate(resultado.get("messages", [])):
            if hasattr(msg, "content"):
                logger.debug(f"  Mensaje {i}: {str(msg.content)[:200]}...")
    else:
        # Validar que tiene todos los ingredientes
        ingredientes_recibidos = len(respuesta_final.get("ingredientes", []))
        if ingredientes_recibidos != len(ingredientes_raw):
            logger.warning(
                f"⚠️ ADVERTENCIA: Se esperaban {len(ingredientes_raw)} ingredientes, "
                f"pero se recibieron {ingredientes_recibidos}"
            )
            logger.warning(f"   Ingredientes esperados: {ingredientes_raw}")
            ingredientes_dict = {ing.get("ingrediente", ""): i for i, ing in enumerate(respuesta_final.get("ingredientes", []))}
            logger.warning(f"   Ingredientes recibidos: {list(ingredientes_dict.keys())}")

    # Validar y convertir a modelo Pydantic
    try:
        analisis = AnálisisReceta(**respuesta_final)
    except Exception as e:
        logger.error(f"Error validando modelo: {e}")
        analisis = AnálisisReceta(
            nombre_receta=respuesta_final.get("nombre_receta", "Receta"),
            ingredientes=[
                IngredienteNutricional(
                    ingrediente=ing.get("ingrediente", ""),
                    cantidad=ing.get("cantidad", ""),
                    calorias=ing.get("calorias", 0),
                    grasas_gramos=ing.get("grasas_gramos", 0)
                )
                for ing in respuesta_final.get("ingredientes", [])
            ],
            totales=TotalesNutricionales(
                total_calorias=respuesta_final.get("totales", {}).get("total_calorias", 0),
                total_grasas=respuesta_final.get("totales", {}).get("total_grasas", 0),
                cantidad_ingredientes=respuesta_final.get("totales", {}).get("cantidad_ingredientes", 0)
            )
        )

    # Mostrar resultados en tabla
    table = Table(title=f"📊 Análisis Nutricional: {analisis.nombre_receta}")
    table.add_column("Ingrediente", style="cyan")
    table.add_column("Cantidad", style="magenta")
    table.add_column("Calorías", justify="right", style="green")
    table.add_column("Grasas (g)", justify="right", style="yellow")

    for ing in analisis.ingredientes:
        table.add_row(
            ing.ingrediente,
            ing.cantidad,
            f"{ing.calorias:.1f}",
            f"{ing.grasas_gramos:.1f}"
        )

    table.add_row(
        "[bold]TOTAL[/bold]",
        "",
        f"[bold green]{analisis.totales.total_calorias:.1f}[/bold green]",
        f"[bold yellow]{analisis.totales.total_grasas:.1f}[/bold yellow]"
    )

    rprint(table)

    # Generar Excel si se solicita
    if guardar_excel:
        archivo_excel = generar_excel_desde_analisis(analisis)
        rprint(f"\n[bold green]✅ Archivo Excel guardado: {archivo_excel}[/bold green]")

    return analisis


# ============================================================================
# 7. EJECUCIÓN CON EJEMPLO
# ============================================================================

if __name__ == "__main__":
    receta_ejemplo = """
    Ensalada Verde Saludable:
    - 200g de lechuga fresca
    - 100g de tomate
    - 50g de pepino
    - 100g de pollo hervido
    - 30g de aceite de oliva
    - 10g de sal y pimienta
    """

    analisis_resultado = procesar_receta(receta_ejemplo, guardar_excel=True)

    rprint("\n[bold blue]📋 JSON Estructurado Completo:[/bold blue]")
    rprint(Pretty(analisis_resultado.model_dump()), "\n")
